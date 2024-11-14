import openpyxl
from bardapi import Bard
import os
import time
import xlwings as xw
import requests
from email import policy
import pickle
import re
from email.parser import BytesParser
from openpyxl.reader.excel import load_workbook



os.environ['_BARD_API_KEY'] = 'XQijNbCRvCgXIDvy9_uSLU4FjYSW0HgEqyRFF0R4-WBKc-2khha8Q4HghzjUBo_eqPAh3g.'
token = 'XQijNbCRvCgXIDvy9_uSLU4FjYSW0HgEqyRFF0R4-WBKc-2khha8Q4HghzjUBo_eqPAh3g.'
def promptings(r):


    if prompt1[r] != None:

        response = bard.get_answer(prompt1[r])['content']
        time.sleep(20)
        ws.cell(row=max_row + 1, column=3).value = response

        if prompt2[r] != None:
            # que3= v6b[r]+ "is a trip. Please retrieve this trip in the form (Traveler-name, departure-date, departure-time, departure-location, arrival-date, arrival-time, arrival-location)"

            response = bard.get_answer(prompt2[r])['content']
            time.sleep(20)
            ws.cell(row=max_row + 1, column=4).value = response

            if prompt3[r] != None:
                response = bard.get_answer(prompt3[r])['content']
                time.sleep(20)
                ws.cell(row=max_row + 1, column=5).value = response



def testemails_bard():
    ws6 = xw.Book("/Users/aishwaryavijayan/Documents/bard/newtrips-annotat.xlsx").sheets['Testemails_trip']
    # ws6 = xw.Book("newtrips-annotat.xlsx").sheets['training']
    # v6b = ws6.range('B3').value
    emails1 = ws6.range("A1:A22").value
    # print(v6b)
    # prompt1 = ws5.range("D2:D40").value
    n1 = 0
    try:
        wb1 = openpyxl.load_workbook('/Users/aishwaryavijayan/aftertraininginbardapi-jun4.xlsx')
    except FileNotFoundError:
        wb1 = openpyxl.Workbook()

    if '1st' in wb1.sheetnames:
        ws1 = wb1['1st']
    else:
        ws1 = wb1.create_sheet('1st')
    objects = []
    with (open("testset-set2.pkl", "rb")) as openfile:
        while True:
            try:
                objects.append(pickle.load(openfile))
            except EOFError:
                break
    ' '.join(str(e) for e in objects)
    fname= None
    #bot = ChatGPT()
    
    
    for fname in objects[0]:

    #for fname in emails1:
        # time.sleep(60)
        n1 = n1 + 1
        if os.path.isfile(user_input + os.sep + str(fname)):
            print('inside')
            # emailset.remove(fname)
            with open(user_input + os.sep + str(fname), 'rb') as f:
                name = f.name  # Get file name
                msg = BytesParser(policy=policy.default).parse(f)
                if (msg.get_body(preferencelist=('plain'))):
                    text = msg.get_body(preferencelist=('plain')).get_content()

                    text = textcleaning(text)
                    newmsg1 = ' '

                    s1 = 'Act as if you are a data annotator. Retrieve any kind of travel info in this text and if possible please convert it to a tuple of format(Traveler-name, departure-date, departure-time, departure-location, arrival-date, arrival-time, arrival-location) in the above text. If some attributes are not known, insert null values. If there is no such tuple found in the text, please respond as NULL. I donot want any sentence to be responded'
                    newmsg1 = text + s1
                    response = bard.get_answer(newmsg1)['content']
                    time.sleep(10)
                    max_row = ws1.max_row

                    # Write data to the next row

                    # ws1.cell(row=max_row + 1, column=1).value = newmsg1
                    ws1.cell(row=max_row + 1, column=1).value = fname
                    ws1.cell(row=max_row + 1, column=2).value = response

                    path = '/Users/aishwaryavijayan/aftertraininginbardapi-jun4.xlsx'
                    wb1.save(path)

    print('tested files')
    print(n1)
def textcleaning(text):
    ###tokenising the email text
# define punctuation
    punctuations = '''!()-[]{};:'"\<>/?@#$%^&*_~'''


# Replace all occurrences of character s with an empty string

    text = re.sub('http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\(\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+', '', text, flags=re.MULTILINE)
    text = text.replace("httpgetwrap"," ")
    text= text.replace(" Sec. ", " ")
    text= text.replace(" Rep. ", " ")
    text= text.replace(" Sen. ", " ")
    text= text.replace(" U.S ", "US")
    text = re.sub('\s+',' ',text)
#text = text.replace("today",msg['Date'])
#print(text)
# remove punctuation from the string
    no_punct = ""
    for char in text:
        if char not in punctuations:
            no_punct = no_punct + char
    return no_punct



session = requests.Session()
session.headers = {
            "Host": "bard.google.com",
            "X-Same-Domain": "1",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.114 Safari/537.36",
            "Content-Type": "application/x-www-form-urlencoded;charset=UTF-8",
            "Origin": "https://bard.google.com",
            "Referer": "https://bard.google.com/",
        }
session.cookies.set("__Secure-1PSID", os.getenv("_BARD_API_KEY")) 
# session.cookies.set("__Secure-1PSID", token) 




n=0
# Open the existing Excel file (or create a new one if it doesn't exist)
try:
    wb = openpyxl.load_workbook('/Users/aishwaryavijayan/Responses_bardapi-jun4.xlsx')
except FileNotFoundError:
    wb = openpyxl.Workbook()

# Select the first worksheet (or create a new one if it doesn't exist)
if 'promptresponses' in wb.sheetnames:
    ws = wb['promptresponses']
else:
    ws = wb.create_sheet('promptresponses')

#ws5 = xw.Book("newtrips-annotat.xlsx").sheets['bard_2ndrun_2ndround']
ws5 = xw.Book("/Users/aishwaryavijayan/Documents/bard/newtrips-annotat.xlsx").sheets['training']
#ws6 = xw.Book("newtrips-annotat.xlsx").sheets['training']
#v6b = ws6.range('B3').value
emails = ws5.range("B2:B20").value
#print(v6b)
prompt1 = ws5.range("D2:D20").value
prompt2 = ws5.range("E2:E20").value
prompt3 = ws5.range("F2:F20").value




bard = Bard(token=token, session=session, timeout=30)
bard = Bard(timeout=10)
user_input = '/Users/aishwaryavijayan/Documents/emails-podesta'
directory = os.listdir(user_input)
visited=[]
n=0
r=0
fname= None
for fname in emails:
        #time.sleep(60)
    print(fname, flush=True)
    txt = ' '
    text= ' '
    print('here')
    text_content = ' '
    text_content1 = ' '
    text_content2 = ' '
    text_content3 = ' '
    text_content4 = ' '
    if os.path.isfile(user_input + os.sep + fname):
        print('inside')
            #emailset.remove(fname)
        with open(user_input + os.sep + fname, 'rb') as f:
                #print('inside')
            name = f.name  # Get file name
            msg = BytesParser(policy=policy.default).parse(f)
            if(msg.get_body(preferencelist=('plain'))):
                txt = msg.get_body(preferencelist=('plain')).get_content()
                    #text=text.replace('\n', '')
                text = textcleaning(txt)
                newmsg = ' '
                s1 = 'Act as if you are a data annotator. Retrieve any kind of travel info in this text and if possible please convert it to a tuple of format (Traveler-name, departure-date, departure-time, departure-location, arrival-date, arrival-time, arrival-location) in the above text. If some attributes are not known, insert null values. If there is no such tuple found in the text, please respond as NULL. I donot want any sentence to be responded'

                newmsg = text + ' ' + s1
                length = len(newmsg);
                print(length)
                    
                    
                if length < 9000:
                    n= n+1

                    response = bard.get_answer(newmsg)['content']
                    time.sleep(10)
                        # Continued conversation without set new session
                        #print(bard.get_answer("Great! Please retrieve such trips in future texts")['content'])
                    max_row = ws.max_row
                    ws.cell(row=max_row + 1, column=1).value = fname
                    ws.cell(row=max_row + 1, column=2).value = response

                    promptings(r)

                else:
                    mystr = ' '
                    part_length = len(text) // 4

                    parts = [text[i:i + part_length] for i in range(0, len(text), part_length)]

                    newmsg = ' '
                    newmsg = parts[0] + s1
                    response = bard.get_answer(newmsg)['content']
                    time.sleep(10)
                    max_row = ws.max_row
                    ws.cell(row=max_row + 1, column=1).value = fname
                    ws.cell(row=max_row + 1, column=2).value = response

                    promptings(r)

                    newmsg = ' '
                    newmsg = parts[1] + s1
                    response = bard.get_answer(newmsg)['content']
                    time.sleep(10)
                    max_row = ws.max_row
                    ws.cell(row=max_row + 1, column=1).value = fname
                    ws.cell(row=max_row + 1, column=2).value = response

                    promptings(r)

                    newmsg = ' '
                    newmsg = parts[2] + s1
                    response = bard.get_answer(newmsg)['content']
                    time.sleep(10)
                    max_row = ws.max_row
                    ws.cell(row=max_row + 1, column=1).value = fname
                    ws.cell(row=max_row + 1, column=2).value = response

                    promptings(r)
                    newmsg = ' '
                    newmsg = parts[3] + s1
                    response = bard.get_answer(newmsg)['content']
                    time.sleep(10)

                    max_row = ws.max_row
                    ws.cell(row=max_row + 1, column=1).value = fname
                    ws.cell(row=max_row + 1, column=2).value = response

                    promptings(r)
                r=r+1
                path = '/Users/aishwaryavijayan/Responses_bardapi-jun4.xlsx'
                wb.save(path)


print('number of trainedfiles=')
print(n)

    #call code to do testing
testemails_bard()
