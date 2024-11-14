from openpyxl.reader.excel import load_workbook
from selenium.webdriver.common.by import By
import xlwings as xw
from seleniumbase import SB
import time
import re
import openpyxl
import os
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from email import policy
from email.parser import BytesParser
from google.cloud import dialogflow_v2beta1 as dialogflow
def textcleaning(text):
    
    ###tokenising the email text
# define punctuation
    punctuations = '''!()-[]{};:'"\<>/?@#$%^&*_~'''


# Replace all occurrences of character s with an empty string

    text = re.sub('http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\(\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+', '', text, flags=re.MULTILINE)
    text = text.replace("httpgetwrap"," ")
    text= text.replace(" Sec. ", " ")
    text= text.replace(" Rep. ", " ")
    text= text.replace(" Sen. ", " ")
    text= text.replace(" U.S ", "US")
    text = re.sub('\s+',' ',text)
#text = text.replace("today",msg['Date'])
#print(text)
# remove punctuation from the string
    no_punct = ""
    for char in text:
        if char not in punctuations:
            no_punct = no_punct + char
    return no_punct

# Open the existing Excel file (or create a new one if it doesn't exist)
try:
    wb = openpyxl.load_workbook('/Users/aishwaryavijayan/table_13may.xlsx')
except FileNotFoundError:
    wb = openpyxl.Workbook()

# Select the first worksheet (or create a new one if it doesn't exist)
if '1st' in wb.sheetnames:
    ws = wb['1st']
else:
    ws = wb.create_sheet('1st')

#ws5 = xw.Book("newtrips-annotat.xlsx").sheets['bard_2ndrun_2ndround']
#ws5 = xw.Book("newtrips-annotat.xlsx").sheets['emailnames']
#v5 = ws5.range("A1:A111").value
#v5 = ws5.range("B1:B51").value
v5='10.eml'
####trying to read only those which didnt give good results
#v5 = ws5.range("A1:A53").value
#v5 = '10372.eml'
#print("Result:", len(v5))
# Set the starting row to the next empty row in the worksheet
start_row = ws.max_row + 1
with SB(uc=True, headed=True) as driver:
    driver.get(
        "https://accounts.google.com/o/oauth2/v2/auth/oauthchooseaccount?redirect_uri=https%3A%2F%2Fdevelopers.google"
        ".com%2Foauthplayground&prompt=consent&response_type=code&client_id=407408718192.apps.googleusercontent.com"
        "&scope=email&access_type=offline&flowName=GeneralOAuthFlow")
    driver.type("#identifierId", 'vignesh.sankaranarayanan87@gmail.com')
    driver.click("#identifierNext > div > button")

    driver.type("#password > div.aCsJod.oJeWuf > div > div.Xb9hP > input", 'ammuvicky')
    driver.click("#passwordNext > div > button")
    driver.get("https://accounts.google.com/ServiceLogin?passive=1209600&continue=https://bard.google.com"
               "/&followup=https://bard.google.com/&ec=GAZAkgU")

    user_input = '/Users/aishwaryavijayan/Documents/emails-podesta'
    directory = os.listdir(user_input)
    #directory = '/Users/aishwaryavijayan/Documents/emails-podesta'
    #file = open("trips18jan.csv", "a")
    visited=[]
    n=0
    fname= None
    #bot = ChatGPT()
    for fname in v5:
        #time.sleep(60)
        n=n+1
        if os.path.isfile(user_input + os.sep + fname) and (fname not in visited):
            #print('inside')
            #emailset.remove(fname)
            with open(user_input + os.sep + fname, 'rb') as f:
                name = f.name  # Get file name
                msg = BytesParser(policy=policy.default).parse(f)
                if(msg.get_body(preferencelist=('plain'))):
                    text = msg.get_body(preferencelist=('plain')).get_content()
                    #text=text.replace('\n', '')
                    text=textcleaning(text)
                    newmsg = ' '
                    
                    #querystring1 = "I will be giving you some examples of a text and a trip/travel tuple of form (Traveler-name, departure-date, departure-time, departure-location, arrival-date, arrival-time, arrival-location) mentioned in the text. I would like to get all similar trips/ travel tuples from an input text."
                    #querystring2 = "(Megan Rouse, Monday 11/24/2014, evening, null, Saturday 11/29/2014, morning, null) is a trip mentioned in the text “Awesome news Cant wait to see you guys. On Sun, Jun 29, 2014 at 1200 AM, Megan Rouse meganrousegmail.com wrote  We have our plane tickets, we fly in Monday 1124 evening and depart Sat  1129 morning.   xo,  Megan.” " 
                    #querystring3 = " (Tom, 2008-11-27, null, Jerusalem, null, null, Luxor) is another trip tuple mentioned in the following paragraph. Dear Friends, Happy Thanksgiving. This year I find myself away from family and friends but after more than a week in the Middle East I feel closer to the rest of the world. The news of Barack Obamas win has excited people herein Jerusalem on both the Arab and Israeli side of the wall, in Cairo the informal capital of the Arab world and in rural Muslim towns along the Nile River north of the High Dam at Aswan. Children no older than 10 break out into chants of Obama Obama Obama and in one school I visited a banner was made with a photo of Obama where children practiced their English by writing Yes we can and I love you on the paper. A glimmer of change. Now I am on the River Nile cruising to Luxor aboard one of the many boats for this purpose. Luxor is close to Karnak and the Valley of the Kings. The scenery is almost biblical. Lush rice fields, palm date groves and banana groves for a few kilometers or so from the river followed by a stark break to desert and dry cliffs. Periodically there is a village marked by the minaret of the mosque. At night it is lit with a green light and the call of the Hazan to prayer echoes across the Nile. Teen boys work the water in boats fishing. They pound the side of the boats with clubs to scare the fish into their nets. On shore camels are dragging gigantic palm leaves. Shephards prepare sheep and goats for market. The temperature is a cool 80 for this winter but the sun is intense. The Egyptian people are very friendly. Quick with funny jokes or a smile. There is inescapable poverty here but very low crime. The crimes are structural on the people, not by the people. All of this is a reminder of the many things to be thankful for and for me that means you. Happy Thanksgiving. From my iPhone on the River Nile, Tom."
                    #query4= "Now please let me know the similar trips mentioned in the text I input. Please retrieve all trips which have at most 3 null attributes. If there is no such tuple found in the text, please respond as NULL. I donot want any sentence to be responded."
                    #newmsg = querystring1+querystring2+querystring3+query4+text
                    #s1 = 'Retrieve all tuples of the form (Traveler-name, departure-date, departure-time, departure-location, arrival-date, arrival-time, arrival-location) in the above text. If some attributes are not known, insert null values. If there is no such tuple found in the text, please respond as NULL. I donot want any sentence to be responded.'
                    #newmsg = text + s1
                    #parts = text.split(" ")
                    #part1 = parts[:len(parts)//2]
                    #part2 = parts[len(parts)//2:]
                    #newmsg=' '
                    #li = list(s1.split(" "))
   
                    #mystr =  part2 + li###only the first halves
                    #Stores the length of the string  
                    #length = len(text);   
#n determines the variable that divide the string in 'n' equal parts  
                    #n = 4;  
                    #temp = 0;  
                    #chars = int(length/n);  
#Stores the array of string  
                    #equalStr = [];   
#Check whether a string can be divided into n equal parts  
                    #if(length % 4 != 0):  
                     #   print("file"+fname"cannot be split);  
                    #else:  
                    part_length = len(text) // 4

                    parts = [text[i:i+part_length] for i in range(0, len(text), part_length)]

                        #print(parts)  
                    mystr = parts[3]+s1
                        #for x in mystr:
                         #   newmsg += ' '+ x
                #case 1
                    driver.type("textarea[id='mat-input-0']", newmsg)
                    #driver.type("textarea[id='mat-input-0']", mystr)
                    time.sleep(60)
                    driver.click('button[mattooltip="Submit"]')
                    time.sleep(20)

                    model_response = driver.execute_script("return document.querySelector('.model-response-text')")
                    text_content = driver.execute_script("return arguments[0].textContent", model_response)

                    max_row = ws.max_row

                    # Write data to the next row
                    #ws.cell(row=max_row + 1, column=1).value = mystr
                    ws.cell(row=max_row + 1, column=1).value = newmsg
                    ws.cell(row=max_row + 1, column=2).value = text_content


                    
                # Save the Excel file
                    path = '/Users/aishwaryavijayan/table1.xlsx'
                    wb.save(path)
                    driver.get("https://bard.google.com")
