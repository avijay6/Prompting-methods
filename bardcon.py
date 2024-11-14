from openpyxl.reader.excel import load_workbook
from selenium.webdriver.common.by import By
import xlwings as xw
from seleniumbase import SB
import time
import re
import openpyxl
import os
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from email import policy
from email.parser import BytesParser
def textcleaning(text):
    ###tokenising the email text
# define punctuation
    punctuations = '''!()-[]{};:'"\<>/?@#$%^&*_~'''


# Replace all occurrences of character s with an empty string

    text = re.sub('http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\(\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+', '', text, flags=re.MULTILINE)
    text = text.replace("httpgetwrap"," ")
    text= text.replace(" Sec. ", " ")
    text= text.replace(" Rep. ", " ")
    text= text.replace(" Sen. ", " ")
    text= text.replace(" U.S ", "US")
    text = re.sub('\s+',' ',text)
#text = text.replace("today",msg['Date'])
#print(text)
# remove punctuation from the string
    no_punct = ""
    for char in text:
        if char not in punctuations:
            no_punct = no_punct + char
    return no_punct

# Open the existing Excel file (or create a new one if it doesn't exist)
try:
    wb = openpyxl.load_workbook('/Users/aishwaryavijayan/Trips-may16.xlsx')
except FileNotFoundError:
    wb = openpyxl.Workbook()

# Select the first worksheet (or create a new one if it doesn't exist)
if '1st' in wb.sheetnames:
    ws = wb['1st']
else:
    ws = wb.create_sheet('1st')

#ws5 = xw.Book("newtrips-annotat.xlsx").sheets['bard_2ndrun_2ndround']
ws5 = xw.Book("newtrips-annotat.xlsx").sheets['emailnames']
v5 = ws5.range("A1:A109").value
#v5 = ws5.range("B1:B51").value
####trying to read only those which didnt give good results
#v5 = ws5.range("A1:A53").value
#v5 = '10372.eml'
print("Result:", len(v5))
# Set the starting row to the next empty row in the worksheet
start_row = ws.max_row + 1
with SB(uc=True, headed=True) as driver:
    driver.get(
        "https://accounts.google.com/o/oauth2/v2/auth/oauthchooseaccount?redirect_uri=https%3A%2F%2Fdevelopers.google"
        ".com%2Foauthplayground&prompt=consent&response_type=code&client_id=407408718192.apps.googleusercontent.com"
        "&scope=email&access_type=offline&flowName=GeneralOAuthFlow")
    driver.type("#identifierId", 'vignesh.sankaranarayanan87@gmail.com')
    driver.click("#identifierNext > div > button")

    driver.type("#password > div.aCsJod.oJeWuf > div > div.Xb9hP > input", 'ammuvicky')
    driver.click("#passwordNext > div > button")
    driver.get("https://accounts.google.com/ServiceLogin?passive=1209600&continue=https://bard.google.com"
               "/&followup=https://bard.google.com/&ec=GAZAkgU")

    user_input = '/Users/aishwaryavijayan/Documents/emails-podesta'
    directory = os.listdir(user_input)
   
    visited=[]
    n=0
    fname= None
    #bot = ChatGPT()
    for fname in v5:
        #time.sleep(60)
        n=n+1
        if os.path.isfile(user_input + os.sep + fname) and (fname not in visited):
            #print('inside')
            #emailset.remove(fname)
            with open(user_input + os.sep + fname, 'rb') as f:
                name = f.name  # Get file name
                msg = BytesParser(policy=policy.default).parse(f)
                if(msg.get_body(preferencelist=('plain'))):
                    text = msg.get_body(preferencelist=('plain')).get_content()
                    #text=text.replace('\n', '')
                    text=textcleaning(text)
                    newmsg = ' '
                    
                    s1 = 'Retrieve all tuples of the form (Traveler-name, departure-date, departure-time, departure-location, arrival-date, arrival-time, arrival-location) in the above text. If some attributes are not known, insert null values.'
                    newmsg = text + s1
                    
                    part_length = len(text) // 4

                    parts = [text[i:i+part_length] for i in range(0, len(text), part_length)] 
                    mystr = parts[3]+s1
                      
                    driver.type("textarea[id='mat-input-0']", newmsg)
                    #driver.type("textarea[id='mat-input-0']", mystr)
                    time.sleep(60)
                    driver.click('button[mattooltip="Submit"]')
                    time.sleep(20)

                    model_response = driver.execute_script("return document.querySelector('.model-response-text')")
                    text_content = driver.execute_script("return arguments[0].textContent", model_response)

                    max_row = ws.max_row

                    # Write data to the next row
                    #ws.cell(row=max_row + 1, column=1).value = mystr
                    ws.cell(row=max_row + 1, column=1).value = newmsg
                    ws.cell(row=max_row + 1, column=2).value = text_content


                    
                # Save the Excel file
                    path = '/Users/aishwaryavijayan/Results-bard-3rdrun.xlsx'
                    wb.save(path)
                    driver.get("https://bard.google.com")
