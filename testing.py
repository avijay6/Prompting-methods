import pickle
import os
from openpyxl.reader.excel import load_workbook
from selenium.webdriver.common.by import By
import xlwings as xw
from seleniumbase import SB
import time
import pickle
import re
import openpyxl
import os
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from email import policy
from email.parser import BytesParser
import codecs
def textcleaning(text):
    ###tokenising the email text
# define punctuation
    punctuations = '''!()-[]{};:'"\<>/?@#$%^&*_~'''


# Replace all occurrences of character s with an empty string

    text = re.sub('http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\(\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+', '', text, flags=re.MULTILINE)
    text = text.replace("httpgetwrap"," ")
    text= text.replace(" Sec. ", " ")
    text= text.replace(" Rep. ", " ")
    text= text.replace(" Sen. ", " ")
    text= text.replace(" U.S ", "US")
    text = re.sub('\s+',' ',text)
#text = text.replace("today",msg['Date'])
#print(text)
# remove punctuation from the string
    no_punct = ""
    for char in text:
        if char not in punctuations:
            no_punct = no_punct + char
    return no_punct
try:
    wb1 = openpyxl.load_workbook('/Users/aishwaryavijayan/Tripsetbeforetraining-setof61-jun6.xlsx')
except FileNotFoundError:
    wb1 = openpyxl.Workbook()
    
if '1st' in wb1.sheetnames:
    ws1 = wb1['1st']
else:
    ws1 = wb1.create_sheet('1st')
    

objects=[]
with SB(uc=True, headed=True) as driver:
    driver.get(
        "https://accounts.google.com/o/oauth2/v2/auth/oauthchooseaccount?redirect_uri=https%3A%2F%2Fdevelopers.google"
        ".com%2Foauthplayground&prompt=consent&response_type=code&client_id=407408718192.apps.googleusercontent.com"
        "&scope=email&access_type=offline&flowName=GeneralOAuthFlow")
    driver.type("#identifierId", 'vignesh.sankaranarayanan87@gmail.com')
    driver.click("#identifierNext > div > button")

    driver.type("#password > div.aCsJod.oJeWuf > div > div.Xb9hP > input", 'ammuvicky')
    driver.click("#passwordNext > div > button")
    driver.get("https://accounts.google.com/ServiceLogin?passive=1209600&continue=https://bard.google.com"
               "/&followup=https://bard.google.com/&ec=GAZAkgU")

    user_input = '/Users/aishwaryavijayan/Documents/emails-podesta'
    directory = os.listdir(user_input)
    with (open("testsetof61.pkl", "rb")) as openfile:
        while True:
            try:
                objects.append(pickle.load(openfile))
            except EOFError:
                break
    #user_input = '/Users/aishwaryavijayan/Documents/emails-podesta'

    ' '.join(str(e) for e in objects)
    print(len(objects[0]))
    fname= None
    #ws6 = xw.Book("/Users/aishwaryavijayan/Documents/bard/newtrips-annotat.xlsx").sheets['Testemails_trip']
    # ws6 = xw.Book("newtrips-annotat.xlsx").sheets['training']
    # v6b = ws6.range('B3').value
    #emails1 = ws6.range("A1:A15").value
    #bot = ChatGPT()
    counter = 0
    i=267
    #for i in range(len(objects[0])):
    #for fname in emails1:
        #fname=objects[0][i]
    #for fname in range(267,len(objects[0])-1):
    for fname in objects[0]:
        #time.sleep(60)
    #n1=n1+1
        if os.path.isfile(user_input + os.sep + str(fname)):
            print(fname)
            #emailset.remove(fname)
            with open(user_input + os.sep + str(fname), 'rb') as f:
                name = f.name  # Get file name
                msg = BytesParser(policy=policy.default).parse(f)
                if(msg.get_body(preferencelist=('plain'))):
                    text = msg.get_body(preferencelist=('plain')).get_content()
                    
                    text=textcleaning(text)
                    newmsg1 = ' '
                    
                    #s1 = 'Act as if you are an NLP expert. Retrieve any kind of travel info in this text and if possible please convert it to a tuple of format (Traveler-name, departure-date, departure-time, departure-location, arrival-date, arrival-time, arrival-location) in the above text. If some attributes are not known, insert null values. If there is no such tuple found in the text, please respond as NULL. I donot want any sentence to be responded'               

                    s1 = 'Act as if you are expert in Named entity recognition. Retrieve any kind of travel info in this text and if possible please convert it to a tuple of format(Traveler-name, departure-date, departure-time, departure-location, arrival-date, arrival-time, arrival-location). If some attributes are not known, insert null values.'
                    newmsg1 = text + s1
                    driver.type("textarea[id='mat-input-0']", newmsg1)
                    
                    time.sleep(60)
                    driver.click('button[mattooltip="Submit"]')
                    time.sleep(20)

                    model_response6 = driver.execute_script("return document.querySelector('.model-response-text')")
                    text_content6 = driver.execute_script("return arguments[0].textContent", model_response6)

                    max_row = ws1.max_row

                    # Write data to the next row
                    
                    #ws1.cell(row=max_row + 1, column=1).value = newmsg1
                    ws1.cell(row=max_row + 1, column=2).value = fname
                    ws1.cell(row=max_row + 1, column=3).value = text_content6
            
                    path = '/Users/aishwaryavijayan/Tripsetbeforetraining-setof61-jun6.xlsx'
                    wb1.save(path)
                    counter = counter + 1
                    driver.get("https://bard.google.com")
                    
