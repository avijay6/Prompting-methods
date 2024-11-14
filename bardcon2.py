
from openpyxl.reader.excel import load_workbook
from selenium.webdriver.common.by import By
import xlwings as xw
from seleniumbase import SB
import time
import openpyxl
import os
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from email import policy
from email.parser import BytesParser

# Open the existing Excel file (or create a new one if it doesn't exist)
try:
    wb = openpyxl.load_workbook('/Users/aishwaryavijayan/Results.xlsx')
except FileNotFoundError:
    wb = openpyxl.Workbook()

# Select the first worksheet (or create a new one if it doesn't exist)
if 'part3of4' in wb.sheetnames:
    ws = wb['part3of4']
else:
    ws = wb.create_sheet('part3of4')

#ws5 = xw.Book("newtrips-annotat.xlsx").sheets['emailnames']
ws5 = xw.Book("newtrips-annotat.xlsx").sheets['partsof4-1']
#v5 = ws5.range("A62:A111").value

####trying to read only those which didnt give good results
v5 = ws5.range("A1:A42").value
print("Result:", len(v5))
# Set the starting row to the next empty row in the worksheet
start_row = ws.max_row + 1
with SB(uc=True, headed=True) as driver:
    driver.get(
        "https://accounts.google.com/o/oauth2/v2/auth/oauthchooseaccount?redirect_uri=https%3A%2F%2Fdevelopers.google"
        ".com%2Foauthplayground&prompt=consent&response_type=code&client_id=407408718192.apps.googleusercontent.com"
        "&scope=email&access_type=offline&flowName=GeneralOAuthFlow")
    driver.type("#identifierId", 'aishwaryav88@gmail.com')
    driver.click("#identifierNext > div > button")

    driver.type("#password > div.aCsJod.oJeWuf > div > div.Xb9hP > input", 'Rudhra16july')
    driver.click("#passwordNext > div > button")
    driver.get("https://accounts.google.com/ServiceLogin?passive=1209600&continue=https://bard.google.com"
               "/&followup=https://bard.google.com/&ec=GAZAkgU")

    user_input = '/Users/aishwaryavijayan/Documents/emails-podesta'
    directory = os.listdir(user_input)
    #directory = '/Users/aishwaryavijayan/Documents/emails-podesta'
    #file = open("trips18jan.csv", "a")
    visited=[]
    n=0
    fname= None
    #bot = ChatGPT()
    for fname in v5:
        #time.sleep(60)
        n=n+1
        if os.path.isfile(user_input + os.sep + fname) and (fname not in visited):
            #print('inside')
            #emailset.remove(fname)
            with open(user_input + os.sep + fname, 'rb') as f:
                name = f.name  # Get file name
                msg = BytesParser(policy=policy.default).parse(f)
                if(msg.get_body(preferencelist=('plain'))):
                    text = msg.get_body(preferencelist=('plain')).get_content()
                    text=text.replace('\n', '')
                    s1 = 'Retrieve all tuples of the form (Traveler-name, departure-date, departure-time, departure-location, arrival-date, arrival-time, arrival-location) in the above text. If some attributes are not known, insert null values. If there is no such tuple found in the text, please respond as NULL. I donot want any sentence to be responded.'
                    #newmsg = text + s1
                    #parts = text.split(" ")
                    #part1 = parts[:len(parts)//2]
                    #part2 = parts[len(parts)//2:]
                    #newmsg=' '
                    #li = list(s1.split(" "))
   
                    #mystr =  part2 + li###only the first halves
                    #Stores the length of the string  
                    length = len(text);   
#n determines the variable that divide the string in 'n' equal parts  
                    #n = 4;  
                    #temp = 0;  
                    #chars = int(length/n);  
#Stores the array of string  
                    #equalStr = [];   
#Check whether a string can be divided into n equal parts  
                    #if(length % 4 != 0):  
                     #   print("file"+fname"cannot be split);  
                    #else:  
                    part_length = len(text) // 4

                    parts = [text[i:i+part_length] for i in range(0, len(text), part_length)]

                        #print(parts)
                   
                    mystr = parts[2]+s1
                        #for x in mystr:
                         #   newmsg += ' '+ x
                    #print(mystr)
                    time.sleep(60)
                #case 1
                    driver.type("textarea[id='mat-input-0']", mystr)
                    time.sleep(80)
                    driver.click('button[mattooltip="Submit"]')
                    time.sleep(20)

                    model_response = driver.execute_script("return document.querySelector('.model-response-text')")
                    text_content = driver.execute_script("return arguments[0].textContent", model_response)

                    max_row = ws.max_row

                    # Write data to the next row
                    ws.cell(row=max_row + 1, column=1).value = mystr
                    ws.cell(row=max_row + 1, column=2).value = text_content


                    
                # Save the Excel file
                    path = '/Users/aishwaryavijayan/Results-bardcon2apr27.xlsx'
                    wb.save(path)
                    driver.get("https://bard.google.com")
