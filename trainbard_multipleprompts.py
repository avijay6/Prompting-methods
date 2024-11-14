from openpyxl.reader.excel import load_workbook
from selenium.webdriver.common.by import By
import xlwings as xw
from seleniumbase import SB
import time
import pickle
import re
import openpyxl
import os
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from email import policy
from email.parser import BytesParser
def testemails_bard(session_id, counter):
    ws6 = xw.Book("/Users/aishwaryavijayan/Documents/bard/newtrips-annotat.xlsx").sheets['Testemails_trip']
    # ws6 = xw.Book("newtrips-annotat.xlsx").sheets['training']
    # v6b = ws6.range('B3').value
    emails1 = ws6.range("A1:A22").value
    # print(v6b)
    #prompt1 = ws5.range("D2:D40").value
    n1=0
    try:
        wb1 = openpyxl.load_workbook('/Users/aishwaryavijayan/Trips-aftertesting-may31.xlsx')
    except FileNotFoundError:
        wb1 = openpyxl.Workbook()
    
    if '1st' in wb1.sheetnames:
        ws1 = wb1['1st']
    else:
        ws1 = wb1.create_sheet('1st')
    
    objects = []
    with (open("testset-set2.pkl", "rb")) as openfile:
        while True:
            try:
                objects.append(pickle.load(openfile))
            except EOFError:
                break
    ' '.join(str(e) for e in objects)
    fname= None
    #bot = ChatGPT()
    #for fname in objects[0]:
    for fname in emails1:
        #time.sleep(60)
        n1=n1+1
        if os.path.isfile(user_input + os.sep + str(fname)):
            print('inside')
            #emailset.remove(fname)
            with open(user_input + os.sep + str(fname), 'rb') as f:
                name = f.name  # Get file name
                msg = BytesParser(policy=policy.default).parse(f)
                if(msg.get_body(preferencelist=('plain'))):
                    text = msg.get_body(preferencelist=('plain')).get_content()
                    
                    text=textcleaning(text)
                    newmsg1 = ' '
                    
                    s1 = 'Act as if you are an NLP expert. Retrieve any kind of travel info in this text and if possible please convert it to a tuple of format(Traveler-name, departure-date, departure-time, departure-location, arrival-date, arrival-time, arrival-location) in the above text. If some attributes are not known, insert null values. If there is no such tuple found in the text, please respond as NULL. I donot want any sentence to be responded'
                    newmsg1 = text + s1
                    driver.type("textarea[id='mat-input-0']", newmsg1)
                    
                    time.sleep(10)
                    driver.click('button[mattooltip="Submit"]')
                    time.sleep(10)

                    model_response6 = driver.execute_script("return document.querySelectorAll('.model-response-text')")
                    count = len(model_response6)
                    while (count - 1) != counter:
                        print("The count does not match the expected count.")
                        model_response6 = driver.execute_script(
                            "return document.querySelectorAll('.model-response-text')")
                        count = len(model_response6)
                        if (count - 1) != counter:
                            #driver.get("https://bard.google.com")
                            timer.sleep(20)
                            #counter = 0
                            #driver.type("textarea[id='mat-input-0']", newmsg1)
                            #time.sleep(10)
                            #driver.click('button[mattooltip="Submit"]')
                            #time.sleep(10)
                            model_response6 = driver.execute_script(
                                "return document.querySelectorAll('.model-response-text')")
                        count = len(model_response6)
                    text_content6 = model_response6[counter].text

                    max_row = ws1.max_row

                    # Write data to the next row
                    
                    #ws1.cell(row=max_row + 1, column=1).value = newmsg1
                    ws1.cell(row=max_row + 1, column=1).value = fname
                    ws1.cell(row=max_row + 1, column=2).value = text_content6
            
    
                    path = '/Users/aishwaryavijayan/Trips-aftertesting-may31.xlsx'
                    wb1.save(path) 

    print('tested files')
    print(n1)
def promptings(r,counter, count, session_id):


    if prompt1[r] != None:
        driver.type("textarea[id='mat-input-0']", prompt1[r])
        time.sleep(10)
        driver.click('button[mattooltip="Submit"]')
        time.sleep(10)
        model_response1 = driver.execute_script("return document.querySelectorAll('.model-response-text')")
        count = len(model_response1)
        while (count - 1) != counter:
            print("The count does not match the expected count.")
            model_response1 = driver.execute_script(
                "return document.querySelectorAll('.model-response-text')")
            count = len(model_response1)
            if (count - 1) != counter:
                #driver.get("https://bard.google.com")
                time.sleep(20)
                #counter = 0
                #if (session_id != driver.driver.session_id):
                 #   print("sessionid doesnt match in testing")
                  #  print(f"old sessionid {session_id}")
                   # session_id = driver.driver.session_id
                   # print(f"new sessionid {session_id}")
                   # exit()
                #session_id = driver.driver.session_id
                #driver.type("textarea[id='mat-input-0']", prompt1[r])
                #time.sleep(10)
                #driver.click('button[mattooltip="Submit"]')
                #time.sleep(10)
                model_response1 = driver.execute_script(
                    "return document.querySelectorAll('.model-response-text')")
                count = len(model_response1)
        text_content1 = model_response1[counter].text
        counter = counter + 1

        ws.cell(row=max_row + 1, column=3).value = text_content1

        if prompt2[r] != None:
            # que3= v6b[r]+ "is a trip. Please retrieve this trip in the form (Traveler-name, departure-date, departure-time, departure-location, arrival-date, arrival-time, arrival-location)"
            driver.type("textarea[id='mat-input-0']", prompt2[r])
            time.sleep(10)
            driver.click('button[mattooltip="Submit"]')
            time.sleep(10)
            model_response2 = driver.execute_script("return document.querySelectorAll('.model-response-text')")
            count = len(model_response2)
            while (count - 1) != counter:
                print("The count does not match the expected count.")
                model_response2 = driver.execute_script(
                    "return document.querySelectorAll('.model-response-text')")
                count = len(model_response2)
                if (count - 1) != counter:
                    #driver.get("https://bard.google.com")
                    time.sleep(20)
                    #counter = 0
                    #if (session_id != driver.driver.session_id):
                     #   print("sessionid doesnt match in testing")
                      #  print(f"old sessionid {session_id}")
                       # session_id = driver.driver.session_id
                        #print(f"new sessionid {session_id}")
                        #exit()
                    #session_id = driver.driver.session_id
                    #driver.type("textarea[id='mat-input-0']", prompt2[r])
                    #time.sleep(10)
                    #driver.click('button[mattooltip="Submit"]')
                    #time.sleep(10)
                    model_response2 = driver.execute_script(
                        "return document.querySelectorAll('.model-response-text')")
                    count = len(model_response2)
            text_content2 = model_response2[counter].text
            counter = counter + 1

            ws.cell(row=max_row + 1, column=4).value = text_content2

            if prompt3[r] != None:
                driver.type("textarea[id='mat-input-0']", prompt3[r])
                time.sleep(10)
                driver.click('button[mattooltip="Submit"]')
                time.sleep(10)
                model_response3 = driver.execute_script("return document.querySelectorAll('.model-response-text')")
                count = len(model_response3)
                while (count - 1) != counter:
                    print("The count does not match the expected count.")
                    model_response3 = driver.execute_script(
                        "return document.querySelectorAll('.model-response-text')")
                    count = len(model_response3)
                    if (count - 1) != counter:
                        #driver.get("https://bard.google.com")
                        time.sleep(20)
                        #counter = 0
                        #if (session_id != driver.driver.session_id):
                         #   print("sessionid doesnt match in testing")
                          ##  print(f"old sessionid {session_id}")
                           # session_id = driver.driver.session_id
                           # print(f"new sessionid {session_id}")
                           # exit()
                       # session_id = driver.driver.session_id
                       # driver.type("textarea[id='mat-input-0']", prompt3[r])
                        #time.sleep(10)
                        #driver.click('button[mattooltip="Submit"]')
                        #time.sleep(10)
                        model_response3 = driver.execute_script(
                            "return document.querySelectorAll('.model-response-text')")
                        count = len(model_response3)
                text_content3 = model_response3[counter].text
                counter = counter + 1

                # ws.cell(row=max_row + 1, column=1).value = fname
                ws.cell(row=max_row + 1, column=5).value = text_content3


def textcleaning(text):
    ###tokenising the email text
# define punctuation
    punctuations = '''!()-[]{};:'"\<>/?@#$%^&*_~'''


# Replace all occurrences of character s with an empty string

    text = re.sub('http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\(\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+', '', text, flags=re.MULTILINE)
    text = text.replace("httpgetwrap"," ")
    text= text.replace(" Sec. ", " ")
    text= text.replace(" Rep. ", " ")
    text= text.replace(" Sen. ", " ")
    text= text.replace(" U.S ", "US")
    text = re.sub('\s+',' ',text)
#text = text.replace("today",msg['Date'])
#print(text)
# remove punctuation from the string
    no_punct = ""
    for char in text:
        if char not in punctuations:
            no_punct = no_punct + char
    return no_punct

n=0
# Open the existing Excel file (or create a new one if it doesn't exist)
try:
    wb = openpyxl.load_workbook('/Users/aishwaryavijayan/Responses-diff_prompts-May31-2.xlsx')
except FileNotFoundError:
    wb = openpyxl.Workbook()

# Select the first worksheet (or create a new one if it doesn't exist)
if 'promptresponses' in wb.sheetnames:
    ws = wb['promptresponses']
else:
    ws = wb.create_sheet('promptresponses')

#ws5 = xw.Book("newtrips-annotat.xlsx").sheets['bard_2ndrun_2ndround']
ws5 = xw.Book("/Users/aishwaryavijayan/Documents/bard/newtrips-annotat.xlsx").sheets['training']
#ws6 = xw.Book("newtrips-annotat.xlsx").sheets['training']
#v6b = ws6.range('B3').value
emails = ws5.range("B2:B40").value
#print(v6b)
prompt1 = ws5.range("D2:D40").value
prompt2 = ws5.range("E2:E40").value
prompt3 = ws5.range("F2:F40").value
print(prompt3)
with SB(uc=True, headed=True) as driver:
    driver.get(
        "https://accounts.google.com/o/oauth2/v2/auth/oauthchooseaccount?redirect_uri=https%3A%2F%2Fdevelopers.google"
        ".com%2Foauthplayground&prompt=consent&response_type=code&client_id=407408718192.apps.googleusercontent.com"
        "&scope=email&access_type=offline&flowName=GeneralOAuthFlow")
    #driver.get(
     #   "https://accounts.google.com/v3/signin/identifier?dsh=S-2045815623%3A1685548963652265&continue=https%3"
      #  "A%2F%2Fbard.google.com%2F&ec=GAZAkgU&ffgf=1&flowEntry=ServiceLogin&flowName=GlifWebSignIn&followup=https%3"
       # "A%2F%2Fbard.google.com%2F&ifkv=Af_xneGNJMc7_xTlrZvbDjgyeJ9YFdlK9k7GRLlScD4K5BQDepIwoW_ZxeYjDzWSVxMGOIoIX6JJmQ"
        #"&passive=1209600"
    #)
    driver.type("#identifierId", 'vignesh.sankaranarayanan87@gmail.com')
    #driver.type("#identifierId", 'aishwaryav88@gmail.com')
    driver.click("#identifierNext > div > button")

    driver.type("#password > div.aCsJod.oJeWuf > div > div.Xb9hP > input", 'ammuvicky')
    #driver.type("#password > div.aCsJod.oJeWuf > div > div.Xb9hP > input", 'Rudhra16july')
    driver.click("#passwordNext > div > button")
    driver.get("https://accounts.google.com/ServiceLogin?passive=1209600&continue=https://bard.google.com"
               "/&followup=https://bard.google.com/&ec=GAZAkgU")

    user_input = '/Users/aishwaryavijayan/Documents/emails-podesta'
    directory = os.listdir(user_input)
    #directory = '/Users/aishwaryavijayan/Documents/emails-podesta'
    #file = open("trips18jan.csv", "a")
    visited=[]
    n=0
    fname= None
    r=0
    counter = 0
    session_id = driver.driver.session_id
    for fname in emails:
        #time.sleep(60)
        print(fname, flush=True)
        if (session_id != driver.driver.session_id):
            print("sessionid doesnt match")
            print(f"old sessionid {session_id}")
            session_id = driver.driver.session_id
            print(f"new sessionid {session_id}")
            exit()
        session_id = driver.driver.session_id
        txt = ' '
        text= ' '
        print('here')
        text_content = ' '
        text_content1 = ' '
        text_content2 = ' '
        text_content3 = ' '
        text_content4 = ' '
        if os.path.isfile(user_input + os.sep + fname):
            print('inside')
            #emailset.remove(fname)
            with open(user_input + os.sep + fname, 'rb') as f:
                #print('inside')
                name = f.name  # Get file name
                msg = BytesParser(policy=policy.default).parse(f)
                if(msg.get_body(preferencelist=('plain'))):
                    txt = msg.get_body(preferencelist=('plain')).get_content()
                    #text=text.replace('\n', '')
                    text = textcleaning(txt)
                    newmsg = ' '
                    s1 = 'Act as if you are an NLP expert. Retrieve any kind of travel info in this text and if possible please convert it to a tuple of format (Traveler-name, departure-date, departure-time, departure-location, arrival-date, arrival-time, arrival-location) in the above text. If some attributes are not known, insert null values. If there is no such tuple found in the text, please respond as NULL. I donot want any sentence to be responded'

                    newmsg = text + s1
                    length = len(newmsg);
                    print(length)
                    
                    
                    if length < 9000:
                        n = n+1

                        driver.type("textarea[id='mat-input-0']", newmsg)

                        time.sleep(30)
                        driver.click('button[mattooltip="Submit"]')
                        time.sleep(20)

                        model_response = driver.execute_script("return document.querySelectorAll('.model-response-text')")
                        count = len(model_response)
                        while (count - 1) != counter:
                            print("The count does not match the expected count.")
                            model_response = driver.execute_script(
                                "return document.querySelectorAll('.model-response-text')")
                            count = len(model_response)
                            if (count - 1) != counter:
                                #driver.get("https://bard.google.com")
                                time.sleep(20)
                                #counter = 0
                                #if (session_id != driver.driver.session_id):
                                 #   print("sessionid doesnt match in testing")
                                  #  print(f"old sessionid {session_id}")
                                   # session_id = driver.driver.session_id
                                   # print(f"new sessionid {session_id}")
                                   # exit()
                                #session_id = driver.driver.session_id
                                #driver.type("textarea[id='mat-input-0']", newmsg)
                                #time.sleep(10)
                                #driver.click('button[mattooltip="Submit"]')
                                #time.sleep(10)
                                model_response = driver.execute_script(
                                    "return document.querySelectorAll('.model-response-text')")
                                count = len(model_response)


                        text_content = model_response[counter].text
                        counter = counter + 1

                        max_row = ws.max_row
                        ws.cell(row=max_row + 1, column=1).value = fname
                        ws.cell(row=max_row + 1, column=2).value = text_content

                        promptings(r, counter, count, session_id)

                
                    else:
                        mystr = ' '
                        part_length = len(text) // 4

                        parts = [text[i:i + part_length] for i in range(0, len(text), part_length)]

                        newmsg = ' '

                        newmsg = parts[0] + s1

                        driver.type("textarea[id='mat-input-0']", newmsg)

                        time.sleep(10)
                        driver.click('button[mattooltip="Submit"]')
                        time.sleep(10)

                        model_response = driver.execute_script("return document.querySelectorAll('.model-response-text')")
                        count= len(model_response)
                        while (count - 1) != counter:
                            print("The count does not match the expected count.")
                            model_response = driver.execute_script(
                                "return document.querySelectorAll('.model-response-text')")
                            count = len(model_response)
                            if (count - 1) != counter:
                                #driver.get("https://bard.google.com")
                                time.sleep(20)
                                #counter = 0
                                #if (session_id != driver.driver.session_id):
                                  ##  print("sessionid doesnt match in testing")
                                 #   print(f"old sessionid {session_id}")
                                    #session_id = driver.driver.session_id
                                    #print(f"new sessionid {session_id}")
                                    #exit()
                               # session_id = driver.driver.session_id
                               # driver.type("textarea[id='mat-input-0']", newmsg)
                               # time.sleep(10)
                                #driver.click('button[mattooltip="Submit"]')
                               # time.sleep(10)
                                model_response = driver.execute_script(
                                    "return document.querySelectorAll('.model-response-text')")
                            count = len(model_response)
                        text_content = model_response[counter].text
                        max_row = ws.max_row
                        ws.cell(row=max_row + 1, column=1).value = fname
                        ws.cell(row=max_row + 1, column=2).value = text_content

                        promptings(r, counter, count, session_id)

                        newmsg = ' '
                        mystr = ' '
                        newmsg = parts[1] + s1

                        driver.type("textarea[id='mat-input-0']", newmsg)

                        time.sleep(10)
                        driver.click('button[mattooltip="Submit"]')
                        time.sleep(10)

                        model_response11 = driver.execute_script("return document.querySelectorAll('.model-response-text')")
                        count = len(model_response11)
                        while (count - 1) != counter:
                            print("The count does not match the expected count.")
                            model_response11 = driver.execute_script(
                                "return document.querySelectorAll('.model-response-text')")
                            count = len(model_response11)
                            if (count - 1) != counter:
                                #driver.get("https://bard.google.com")
                                time.sleep(20)
                                #counter = 0
                                #if (session_id != driver.driver.session_id):
                                 #   print("sessionid doesnt match in testing")
                                  #  print(f"old sessionid {session_id}")
                                   # session_id = driver.driver.session_id
                                   # print(f"new sessionid {session_id}")
                                   # exit()
                                #session_id = driver.driver.session_id
                                #driver.type("textarea[id='mat-input-0']", newmsg)
                                #time.sleep(10)
                                #driver.click('button[mattooltip="Submit"]')
                                #time.sleep(10)
                                model_response11 = driver.execute_script(
                                    "return document.querySelectorAll('.model-response-text')")
                            count = len(model_response11)
                        text_content11 = model_response11[counter].text
                        max_row = ws.max_row
                        ws.cell(row=max_row + 1, column=1).value = fname
                        ws.cell(row=max_row + 1, column=2).value = text_content11

                        promptings(r, counter, count, session_id)

                        newmsg = ' '
                        mystr = ' '
                        newmsg = parts[2] + s1

                        driver.type("textarea[id='mat-input-0']", newmsg)

                        time.sleep(10)
                        driver.click('button[mattooltip="Submit"]')
                        time.sleep(10)

                        model_response11 = driver.execute_script(
                            "return document.querySelectorAll('.model-response-text')")
                        count = len(model_response11)
                        while (count - 1) != counter:
                            print("The count does not match the expected count.")
                            model_response11 = driver.execute_script(
                                "return document.querySelectorAll('.model-response-text')")
                            count = len(model_response11)
                            if (count - 1) != counter:
                                #driver.get("https://bard.google.com")
                                time.sleep(20)
                                #counter = 0
                                #if (session_id != driver.driver.session_id):
                                 #   print("sessionid doesnt match in testing")
                                  #  print(f"old sessionid {session_id}")
                                  #  session_id = driver.driver.session_id
                                  #  print(f"new sessionid {session_id}")
                                  #  exit()
                                #session_id = driver.driver.session_id
                                #driver.type("textarea[id='mat-input-0']", newmsg)
                                #time.sleep(10)
                                #driver.click('button[mattooltip="Submit"]')
                                #time.sleep(10)
                                model_response11 = driver.execute_script(
                                    "return document.querySelectorAll('.model-response-text')")
                            count = len(model_response11)
                        text_content11 = model_response11[counter].text

                        max_row = ws.max_row
                        ws.cell(row=max_row + 1, column=1).value = fname
                        ws.cell(row=max_row + 1, column=2).value = text_content11

                        promptings(r, counter, count, session_id)

                        newmsg = ' '
                        mystr = ' '
                        newmsg = parts[3] + s1
                        #for x in newmsg:
                         #   mystr += ' ' + x

                        driver.type("textarea[id='mat-input-0']", newmsg)

                        time.sleep(10)
                        driver.click('button[mattooltip="Submit"]')
                        time.sleep(10)

                        model_response11 = driver.execute_script(
                            "return document.querySelectorAll('.model-response-text')")
                        count = len(model_response11)
                        while (count - 1) != counter:
                            print("The count does not match the expected count.")
                            model_response11 = driver.execute_script(
                                "return document.querySelectorAll('.model-response-text')")
                            count = len(model_response11)
                            if (count - 1) != counter:
                                #driver.get("https://bard.google.com")
                                time.sleep(20)
                                #counter = 0
                                ##if (session_id != driver.driver.session_id):
                                  #  print("sessionid doesnt match in testing")
                                   # print(f"old sessionid {session_id}")
                                   # session_id = driver.driver.session_id
                                   # print(f"new sessionid {session_id}")
                                    #exit()
                               # session_id = driver.driver.session_id
                               # driver.type("textarea[id='mat-input-0']", newmsg)
                               # time.sleep(10)
                               # driver.click('button[mattooltip="Submit"]')
                               # time.sleep(10)
                                model_response11 = driver.execute_script(
                                    "return document.querySelectorAll('.model-response-text')")
                            count = len(model_response11)
                        text_content11 = model_response11[counter].text
                        max_row = ws.max_row
                        ws.cell(row=max_row + 1, column=1).value = fname
                        ws.cell(row=max_row + 1, column=2).value = text_content11

                        promptings(r, counter, count, session_id)
                    r=r+1
                    path = '/Users/aishwaryavijayan/Responses-diff_prompts-May31-2.xlsx'
                    wb.save(path)


    print('number of trainedfiles=')
    print(n)

    #call code to do testing
    testemails_bard(session_id, counter)
    #path = '/Users/aishwaryavijayan/Trips-aftertesting-unseen-emails.xlsx'
    #wb1.save(path)              
                # Save the Excel file
                    
                    #driver.get("https://bard.google.com")


                    
